# encoding: utf-8
import datetime
import openpyxl, os, time

from openpyxl import load_workbook, Workbook

'''
功能：打开一个excel文件
参数：等待打开的excel文件名
返回值：打开的excel文件资源
'''


def openOneExcel(excelName=''):
  try:
    workbook = load_workbook(filename = excelName, read_only=True)
    return workbook
  except Exception as e:
    print(e)
    print(">>>>>>>>>>>>读取[%s]异常<<<<<<<<<<<<" % (excelName))
    return None


'''
功能：将制定目录下的文件名转换成列表
参数：目录名
返回值：当前目录下所有符合的文件
'''


def dir2list(dirName):
  filelists = os.listdir(dirName)
  filelists = [i for i in filelists if i.endswith(".xlsx")];
  filelists = [i for i in filelists if not i.startswith("~")];
  return filelists


'''
功能：读一个excel文件
参数：文件名
返回值：文件下的所有sheet
'''


def readOneExcel(excelName=''):
  book = openOneExcel(excelName)
  if book is None:
    return None
  sheetNames = book.sheetnames
  sheets = []
  for i in sheetNames:
    sheet = book.get_sheet_by_name(i)
    sheets.append(sheet)
  return sheets


'''
功能：读文件，写入excel
参数：目录名，写入的文件名
返回值：无
'''


def writeToOneExcel(dirName='', toExcelName=''):
  filelists = dir2list(dirName)
  rowNum = 1
  colNum = 1
  wb = Workbook()

  # 为工作簿添加sheet1工作表
  ws = wb.active;
  ws.title = "Sheet"

  for afile in filelists:
    starttime = datetime.datetime.now()
    print ("begin [%s] ,files[%s]" % (time.ctime(), afile))
    lists = readOneExcel(dirName + "/" + afile)
    if lists is None:
      return
    for sheet in lists:
      for row in sheet.rows:
        for col in row:
          # if rowNum==1:
          #     ws.column_dimensions[num2column(colNum)].width = 18.0
          try:
            if col.value is not None:
              if col.data_type == 'd':
                ws.cell(row=rowNum, column=colNum).value = col.value
              if col.data_type == 's':
                ws.cell(row=rowNum, column=colNum).value = col.value.encode(
                    'gbk').decode('gbk').encode('utf8')
              if isinstance(col.value, long):
                ws.cell(row=rowNum, column=colNum).value = long(col.value)
              # print(col.value)
          except Exception as e:
            a = e;
          colNum = colNum + 1
        rowNum = rowNum + 1
        colNum = 1
      wb.save(toExcelName)
    endtime = datetime.datetime.now()
    print ("end 用时[%s]秒 ,files[%s]" % ((endtime - starttime).seconds, afile))
    print("")
  print ("合并结束")


# 将列数转成列名对应单元格
def num2column(num):
  interval = ord('Z') - ord('A')
  tmp = ''
  multiple = num // interval
  remainder = num % interval
  while multiple > 0:
    if multiple > 25:
      tmp += 'A'
    else:
      tmp += chr(64 + multiple)
    multiple = multiple // interval
  tmp += chr(64 + remainder)
  return tmp


def main():
  path = os.path.abspath('.')
  dirName = path + "/aba"
  toExcelName = path + "/total1.xlsx"
  writeToOneExcel(dirName, toExcelName)
  # print(num2column(2))


if __name__ == "__main__":
  main()
