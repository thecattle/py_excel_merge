#encoding: utf-8
import threading
from tkinter import scrolledtext, ttk
from tkinter import*
import datetime
import os
import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

'''
功能：打开一个excel文件
参数：等待打开的excel文件名
返回值：打开的excel文件资源
'''


def openOneExcel(excelName=''):
  try:
    workbook = load_workbook(filename = excelName, read_only=True,data_only=True)
    return workbook
  except Exception as e:
    print(e)
    print(">>>>>>>>>>>>读取[%s]异常<<<<<<<<<<<<" % (excelName))
    return None


'''
功能：将制定目录下的文件名转换成列表
参数：目录名
返回值：当前目录下所有符合的文件
'''


def dir2list(dirName):
  filelists = os.listdir(dirName)
  filelists = [i for i in filelists if i.endswith(".xlsx")];
  filelists = [i for i in filelists if not i.startswith("~")];
  return filelists


'''
功能：读一个excel文件
参数：文件名
返回值：文件下的所有sheet
'''


def readOneExcel(excelName=''):
  book = openOneExcel(excelName)
  if book is None:
    return None
  sheetNames = book.sheetnames
  sheets = []
  for i in sheetNames:
    # sheet = book.get_sheet_by_name(book[i])
    sheet = book[i]

    if type == 1:
      sheets.append(sheet)
    else:
      inputSheetKey=entry1.get()
      if inputSheetKey.strip()=='':
        raise Exception("特殊的sheet，得告诉我是啥 比如 明细")
      p1 = entry1.get()
      if p1 in str(i):
        sheets.append(sheet)
  return sheets


'''
功能：读文件，写入excel
参数：目录名，写入的文件名
返回值：无
'''


def writeToOneExcel(dirName='', toExcelName=''):
  filelists = dir2list(dirName)
  if len(filelists):
    printToPanel("一共有%d份excel表格，开始处理" %(len(filelists)))
  else:
    raise Exception("files文件夹内无数据")

  rowNum = 1
  colNum = 1
  wb = Workbook()

  # 为工作簿添加sheet1工作表
  ws = wb.active
  ws.title = "Sheet"

  for afile in filelists:
    starttime = datetime.datetime.now()
    printToPanel ("begin [%s] ,files[%s]" % (time.ctime(), afile))
    lists = readOneExcel(dirName + "/" + afile)
    if lists is None:
      return
    for sheet in lists:
      ws.cell(row=rowNum, column=colNum).value = str(afile)+" - "+sheet.title
      rowNum=rowNum+1
      for row in sheet.rows:
        notBlank=False
        for col in row:
          if rowNum==2:
              ws.column_dimensions[get_column_letter(colNum)].width = 18.0
          try:
            if col.value is not None:
              notBlank=True
              cell=ws.cell(row=rowNum, column=colNum)
              cell.border=col.border
              cell.fill=col.fill
              cell.font=col.font
              cell.alignment=col.alignment
              cell.value = col.value
              print(cell.value)
          except Exception as e:
            print(e)
            printToPanel(str(e))
          colNum = colNum + 1
        if notBlank:
          rowNum = rowNum + 1
        colNum = 1
      #每个sheet空一行
      rowNum=rowNum+1


      wb.save(toExcelName)

    endtime = datetime.datetime.now()
    printToPanel ("end 用时[%s]秒 ,files[%s]" % ((endtime - starttime).seconds, afile))
    printToPanel("")
  print("")

def main():
  try:
    path = os.path.abspath('.')
    dirName=os.path.join(path, 'files')
    # dirName=os.path.join(path, '/Users/sunpeng/Develop/excel/inputs')
    toExcelName = path + "/total.xlsx"
    writeToOneExcel(dirName, toExcelName)
    words = [
      "代码跑完啦，看看对不对",
      "代码跑完啦，还在忙吗",
      "代码跑完啦，还不理我已经过去2分钟了",
      "要多喝水，多去厕所",
      "已经过去4分钟了",
      "已经过去5分钟了",
      "看看我吧，不要总低头看发票，对眼睛不好",
      "我不行了，我困了",
      "emmmm。。。",
      "大河向东流啊 天上的星星参北斗啊  啊啊啊",
      "我是一只小小小小鸟",
      "飞呀飞不高嗷嗷嗷嗷",
      "你背面有个人 长的真帅",
      "拜拜了您呐",
      "拜拜了您呐那哪拿",
    ]
    for i in words:
      printToPanel(i)
      time.sleep(60)
  except Exception as e:
    print(e)
    printToPanel(str(e))
    printToPanel("error 出错啦  ，快截图给我瞅瞅")
    print(type)





###################################################excel#########################################################
myWindow=Tk()
myWindow.title('excel合并')

#设置窗口大小
width = 900
height = 500
type=0

v=IntVar()

#获取屏幕尺寸以计算布局参数，使窗口居屏幕中央
screenwidth = myWindow.winfo_screenwidth()
screenheight = myWindow.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth-width)/2, (screenheight-height)/2)
myWindow.geometry(alignstr)

#设置窗口是否可变长、宽，True：可变，False：不可变
myWindow.resizable(width=False, height=True)
Label(myWindow,text='有问题及时截图给我，xls格式的文件，需要手动保存成xlsx的哦').pack(anchor=W)


def printToPanel(txt):
  scr.insert(END,txt+"\n")
  scr.update()
  scr.see(END)

#定义单选按钮的响应函数
def callRB():
  global type
  type=v.get()

def start():
  if btn.instate(['!disabled']):
    btn.state(['disabled'])
  if type==0:
    printToPanel("单选框得选一个！！！！！！！！")
    raise Exception("单选框得选一个！！！！！！！！")
  if type!=0:
    th=threading.Thread(target=main)
    th.setDaemon(True)#守护线程
    th.start()



btn=ttk.Button(myWindow,text='确定',command=start)
btn.pack(side=BOTTOM)

#列表中存储的是元素是元组
language=[('无',0),('所有文件所有sheet合并',1),('所有文件特殊sheet合并',2)]

scr = scrolledtext.ScrolledText(myWindow, width=70, height=13,font=("隶书",18))  #滚动文本框（宽，高（这里的高应该是以行数为单位），字体样式）
scr.place(x=20, y=130) #滚动文本框在页面的位置

Label(myWindow, text="要单独合并的sheet名/共同的字，例如：明细表，明细表1，明细2。则输入\"明细\" :").place(x=10,y=102)

# #Entry控件布局
entry1=Entry(myWindow)
entry1.place(x=490,y=100)


#for循环创建单选框
for lan,num in language:
  Radiobutton(myWindow, text=lan, value=num, command=callRB, variable=v).pack(anchor=W)


#进入消息循环
myWindow.mainloop()